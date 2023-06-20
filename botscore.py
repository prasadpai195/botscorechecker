# The script helps look through each security configs and policies for action defined on Command injection Attack group and for KRS rule ID 3000014
# Minimal error handling, use at your risk
import requests
import json
from akamai.edgegrid import EdgeGridAuth, EdgeRc
import os
from urllib.parse  import urljoin
from openpyxl import load_workbook
import openpyxl, pprint
import sys
from nested_lookup import nested_lookup
#Enter SwitchKey below

#Enter API creds below
edgerc_file = os.path.join(os.path.expanduser("~"), '.edgerc')
edgerc = EdgeRc(edgerc_file)
section="botscore"
base_url = edgerc.get(section,'host')
baseurl=str('https://')+str(base_url)
client_token=edgerc.get(section,'client_token')
client_secret=edgerc.get(section,'client_secret')
access_token=edgerc.get(section,'access_token')
s = requests.Session()
s.auth = EdgeGridAuth(
client_token=client_token,
client_secret=client_secret,
access_token=access_token
)
if __name__ == '__main__':
    wb = load_workbook('input.xlsx')
    sheet = wb['Sheet1']
    #sheet2=wb['Sheet2']
    print("Number of entries is:" ,sheet.max_row-1, "accounts")
    i=2
    for row in range(2, sheet.max_row+1):
        account=sheet['A' + str(row)].value
        skey=sheet['B' + str(row)].value
        package=sheet['C' + str(row)].value
        print("*************************************************************************************************************************************************************")
        print("Running for "+ account  +"")
        # PULL CONFIGURATIONS
        config=s.get(baseurl + ("/appsec/v1/configs?accountSwitchKey="+skey) , headers = {'PAPI-Use-Prefixes': 'true'})
        configs = json.loads(config.text)
        #print(configs)
        configlist = (configs['configurations'])
        for gdictionary in configlist:
            #RESET to handle exceptions
            prod="None"
            stage="None"
            latest="None"
            for key,value in gdictionary.items():
                configid= str(gdictionary['id'])
                configname=str(gdictionary['name'])
                #print(configname)
                try:
                    prod=str(gdictionary['productionVersion'])
                except(KeyError):
                    key="bypass"    
                try:
                    stage=str(gdictionary['stagingVersion'])
                except(KeyError):
                    key="bypass"
                
                latest=str(gdictionary['latestVersion'])
                
                
            #print("\n Config Name: "+configname)
            if latest != "None":
                prodversion=latest
            if stage != "None":
                prodversion=stage
            if prod != "None":
                prodversion=prod
            else:
                continue;
            policy=s.get(baseurl + ("/appsec/v1/configs/"+configid+"/versions/"+prodversion+"/security-policies?accountSwitchKey="+skey) , headers = {'PAPI-Use-Prefixes': 'true'})
            pol=json.loads(policy.text)
            plist = (pol['policies'])
            for policy in plist:
                pid= str(policy['policyId'])
                policyname= str(policy['policyName'])
                botmanagementenabled=s.get(baseurl + ("/appsec/v1/configs/"+configid+"/versions/"+prodversion+"/security-policies/"+pid+"/bot-management-settings?accountSwitchKey="+skey))
                botenabled=json.loads(botmanagementenabled.text)
                if botenabled["enableBotManagement"] == True:
                	#print("The policy " +policyname+ "has Botmanager enabled, lets check the Botscore here" +configname+"test")
                    botscorecheck=s.get(baseurl + ("/appsec/v1/configs/"+configid+"/versions/"+prodversion+"/security-policies/"+pid+"/transactional-endpoints/bot-protection?accountSwitchKey="+skey))
                    botscore=json.loads(botscorecheck.text)
                    for api in botscore["operations"]:
                        testbsinline=""
                        testbsios=""
                        testbsdroid=""
                        testbsstandard=""
                        nativeSDK=api["telemetryTypeStates"]["nativeSdk"]["enabled"]
                        standard=api["telemetryTypeStates"]["standard"]["enabled"]
                        inline=api["telemetryTypeStates"]["inline"]["enabled"]
                        #print(api["traffic"])
                        if nativeSDK==True:
                            snippetdroid=api["traffic"]["nativeSdkAndroid"]
                            snippetios=api["traffic"]["nativeSdkIos"]
                            datadroid=json.loads(json.dumps(snippetdroid))
                            dataios=json.loads(json.dumps(snippetios))
                            testbsdroid=nested_lookup("aggressiveThreshold",datadroid)
                            testbsios=nested_lookup("aggressiveThreshold",dataios)         
                        if inline==True:
                            snippetinline=api["traffic"]["inlineTelemetry"]
                            datainline=json.loads(json.dumps(snippetinline))
                            testbsinline=nested_lookup("aggressiveThreshold",datainline)
                        if standard==True:
                            snippetstandard=api["traffic"]["standardTelemetry"]
                            datastandard=json.loads(json.dumps(snippetstandard))
                            testbsstandard=nested_lookup("aggressiveThreshold",datastandard)
                        else:
                            continue;
                        if str(testbsdroid) or str(testbsios) or str(testbsinline) or str(testbsstandard) != "[]":
                            print("The policy" +policyname+" and the config "+configname+" has Botscore enabled")
                            break;
                        else:
                            print("The policy"+policyname+" and the config "+configname+"does not have Botscore enabled")